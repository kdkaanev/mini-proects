import openpyxl 
from openpyxl.styles import PatternFill
my_set = set()
search_path = 'b.xlsx'
target_path = 'sliter.xlsx'
wb_obj = openpyxl.load_workbook(search_path) 
sheet_obj = wb_obj.active

for i in range(407):
    cell_obj = sheet_obj.cell(row = i + 1, column = 1)
    my_set.add(str(cell_obj.value))
wb = openpyxl.load_workbook(target_path)
ws = wb['Sheet1'] #Name of the working sheet

for i in range(180):
    cell_obj = ws.cell(row = i + 1, column = 1)
    if cell_obj.value in my_set:
        cell_obj.fill = PatternFill(patternType='solid', 
                           fgColor='7A9174') #green '7A9174'

        wb.save("sliter.xlsx")
